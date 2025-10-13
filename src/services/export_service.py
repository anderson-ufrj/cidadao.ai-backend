"""
Module: services.export_service
Description: Document export service for generating PDF, Excel and CSV files
Author: Anderson H. Silva
Date: 2025-01-25
License: Proprietary - All rights reserved
"""

import asyncio
import io
import zipfile
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from typing import Any, Optional, Union

import markdown
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from src.core import get_logger

logger = get_logger(__name__)

# Thread pool for CPU-intensive PDF generation
_pdf_thread_pool = ThreadPoolExecutor(max_workers=2, thread_name_prefix="pdf_export")


class ExportService:
    """Service for exporting documents in various formats."""

    def __init__(self):
        """Initialize export service."""
        self.styles = getSampleStyleSheet()
        self._create_custom_styles()

    def _create_custom_styles(self):
        """Create custom PDF styles."""
        # Title style
        self.styles.add(
            ParagraphStyle(
                name="CustomTitle",
                parent=self.styles["Title"],
                fontSize=24,
                textColor=colors.HexColor("#1a73e8"),
                spaceAfter=30,
                alignment=TA_CENTER,
            )
        )

        # Subtitle style
        self.styles.add(
            ParagraphStyle(
                name="CustomSubtitle",
                parent=self.styles["Heading2"],
                fontSize=16,
                textColor=colors.HexColor("#34495e"),
                spaceBefore=20,
                spaceAfter=10,
            )
        )

        # Body text style
        self.styles.add(
            ParagraphStyle(
                name="CustomBody",
                parent=self.styles["BodyText"],
                fontSize=11,
                leading=16,
                alignment=TA_JUSTIFY,
                spaceBefore=6,
                spaceAfter=6,
            )
        )

        # Footer style
        self.styles.add(
            ParagraphStyle(
                name="CustomFooter",
                parent=self.styles["Normal"],
                fontSize=9,
                textColor=colors.grey,
                alignment=TA_CENTER,
            )
        )

    async def generate_pdf(
        self,
        content: str,
        title: str,
        metadata: Optional[dict[str, Any]] = None,
        format_type: str = "report",
    ) -> bytes:
        """
        Generate PDF from content.

        Args:
            content: Content in markdown format
            title: Document title
            metadata: Additional metadata
            format_type: Type of document (report, investigation, analysis)

        Returns:
            PDF bytes
        """
        # Run PDF generation in thread pool to avoid blocking
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(
            _pdf_thread_pool,
            self._generate_pdf_sync,
            content,
            title,
            metadata or {},
            format_type,
        )

    def _generate_pdf_sync(
        self, content: str, title: str, metadata: dict[str, Any], format_type: str
    ) -> bytes:
        """Synchronous PDF generation."""
        # Create buffer
        buffer = io.BytesIO()

        # Create PDF
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=48,
        )

        # Build story
        story = []

        # Add header with logo/branding
        story.append(
            Paragraph(
                "Cidadão.AI - Transparência Governamental", self.styles["CustomFooter"]
            )
        )
        story.append(Spacer(1, 0.2 * inch))

        # Add title
        story.append(Paragraph(title, self.styles["CustomTitle"]))

        # Add metadata if provided
        if metadata:
            meta_data = []
            if "generated_at" in metadata:
                meta_data.append(f"Gerado em: {metadata['generated_at']}")
            if "report_type" in metadata:
                meta_data.append(f"Tipo: {metadata['report_type']}")
            if "author" in metadata:
                meta_data.append(f"Autor: {metadata['author']}")

            if meta_data:
                story.append(
                    Paragraph(" | ".join(meta_data), self.styles["CustomFooter"])
                )
                story.append(Spacer(1, 0.3 * inch))

        # Convert markdown to HTML
        html_content = markdown.markdown(
            content, extensions=["extra", "codehilite", "toc", "tables"]
        )

        # Parse HTML and convert to PDF elements
        soup = BeautifulSoup(html_content, "html.parser")

        for element in soup.find_all():
            if element.name == "h1":
                story.append(PageBreak())
                story.append(Paragraph(element.text, self.styles["Heading1"]))
            elif element.name == "h2":
                story.append(Spacer(1, 0.2 * inch))
                story.append(Paragraph(element.text, self.styles["CustomSubtitle"]))
            elif element.name == "h3":
                story.append(Spacer(1, 0.15 * inch))
                story.append(Paragraph(element.text, self.styles["Heading3"]))
            elif element.name == "p":
                story.append(Paragraph(element.text, self.styles["CustomBody"]))
            elif element.name == "ul":
                for li in element.find_all("li"):
                    story.append(Paragraph(f"• {li.text}", self.styles["CustomBody"]))
            elif element.name == "ol":
                for i, li in enumerate(element.find_all("li"), 1):
                    story.append(
                        Paragraph(f"{i}. {li.text}", self.styles["CustomBody"])
                    )
            elif element.name == "table":
                # Convert HTML table to ReportLab table
                table_data = []
                rows = element.find_all("tr")

                for row in rows:
                    cells = row.find_all(["td", "th"])
                    table_data.append([cell.text.strip() for cell in cells])

                if table_data:
                    t = Table(table_data)
                    t.setStyle(
                        TableStyle(
                            [
                                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                ("FONTSIZE", (0, 0), (-1, 0), 12),
                                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                            ]
                        )
                    )
                    story.append(Spacer(1, 0.1 * inch))
                    story.append(t)
                    story.append(Spacer(1, 0.1 * inch))

        # Add footer
        story.append(Spacer(1, 0.5 * inch))
        story.append(
            Paragraph(
                f"Documento gerado automaticamente pelo Cidadão.AI em {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                self.styles["CustomFooter"],
            )
        )

        # Build PDF
        doc.build(story)

        # Get PDF bytes
        pdf_bytes = buffer.getvalue()
        buffer.close()

        return pdf_bytes

    async def generate_excel(
        self,
        data: Union[dict[str, pd.DataFrame], pd.DataFrame],
        title: str,
        metadata: Optional[dict[str, Any]] = None,
    ) -> bytes:
        """
        Generate Excel file from data.

        Args:
            data: DataFrame or dict of DataFrames (for multiple sheets)
            title: Document title
            metadata: Additional metadata

        Returns:
            Excel bytes
        """
        buffer = io.BytesIO()

        # Create Excel writer
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            # Handle single or multiple DataFrames
            if isinstance(data, pd.DataFrame):
                data = {"Dados": data}

            # Write each DataFrame to a sheet
            for sheet_name, df in data.items():
                df.to_excel(
                    writer, sheet_name=sheet_name[:31], index=False
                )  # Excel sheet name limit

                # Get the worksheet
                worksheet = writer.sheets[sheet_name]

                # Apply formatting
                self._format_excel_sheet(worksheet, title, metadata)

            # Add metadata sheet
            if metadata:
                meta_df = pd.DataFrame(
                    [{"Campo": k, "Valor": str(v)} for k, v in metadata.items()]
                )
                meta_df.to_excel(writer, sheet_name="Metadados", index=False)
                self._format_excel_sheet(writer.sheets["Metadados"], "Metadados", {})

        return buffer.getvalue()

    def _format_excel_sheet(self, worksheet, title: str, metadata: dict[str, Any]):
        """Apply formatting to Excel worksheet."""
        # Set column widths
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(
                length + 2, 50
            )

        # Add title row
        worksheet.insert_rows(1)
        worksheet.merge_cells("A1:" + get_column_letter(worksheet.max_column) + "1")
        title_cell = worksheet["A1"]
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True, color="1a73e8")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add generation date
        worksheet.insert_rows(2)
        worksheet.merge_cells("A2:" + get_column_letter(worksheet.max_column) + "2")
        date_cell = worksheet["A2"]
        date_cell.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        date_cell.font = Font(size=10, italic=True)
        date_cell.alignment = Alignment(horizontal="center")

        # Format headers
        header_fill = PatternFill(
            start_color="4285F4", end_color="4285F4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for cell in worksheet[4]:  # Assuming headers are now in row 4
            if cell.value:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

        # Add borders
        from openpyxl.styles import Border, Side

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in worksheet.iter_rows(min_row=4):
            for cell in row:
                if cell.value:
                    cell.border = thin_border

    async def generate_csv(self, data: pd.DataFrame, encoding: str = "utf-8") -> bytes:
        """
        Generate CSV file from DataFrame.

        Args:
            data: DataFrame to export
            encoding: File encoding

        Returns:
            CSV bytes
        """
        return data.to_csv(index=False).encode(encoding)

    async def generate_bulk_export(
        self, exports: list[dict[str, Any]], format: str = "zip"
    ) -> bytes:
        """
        Generate bulk export with multiple files.

        Args:
            exports: List of export configurations
                Each dict should have: 'filename', 'content', 'format'
            format: Archive format (zip)

        Returns:
            Archive bytes
        """
        if format != "zip":
            raise ValueError("Currently only ZIP format is supported for bulk exports")

        buffer = io.BytesIO()

        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
            for export in exports:
                filename = export["filename"]
                content = export["content"]
                file_format = export.get("format", "txt")

                # Generate file based on format
                if file_format == "pdf":
                    file_bytes = await self.generate_pdf(
                        content=content,
                        title=export.get("title", filename),
                        metadata=export.get("metadata", {}),
                    )
                elif file_format == "excel":
                    file_bytes = await self.generate_excel(
                        data=export.get("data", pd.DataFrame()),
                        title=export.get("title", filename),
                        metadata=export.get("metadata", {}),
                    )
                elif file_format == "csv":
                    file_bytes = await self.generate_csv(
                        data=export.get("data", pd.DataFrame())
                    )
                else:
                    # Default to text
                    file_bytes = content.encode("utf-8")

                # Add to zip
                zipf.writestr(filename, file_bytes)

                logger.info(
                    "bulk_export_file_added",
                    filename=filename,
                    format=file_format,
                    size=len(file_bytes),
                )

        return buffer.getvalue()

    async def convert_investigation_to_excel(
        self, investigation_data: dict[str, Any]
    ) -> bytes:
        """
        Convert investigation data to Excel format.

        Args:
            investigation_data: Investigation data dict

        Returns:
            Excel bytes
        """
        # Create multiple DataFrames for different aspects
        dataframes = {}

        # Summary sheet
        summary_data = {
            "Campo": [
                "ID",
                "Tipo",
                "Status",
                "Data Início",
                "Data Fim",
                "Duração (min)",
            ],
            "Valor": [
                investigation_data.get("id", "N/A"),
                investigation_data.get("type", "N/A"),
                investigation_data.get("status", "N/A"),
                investigation_data.get("created_at", "N/A"),
                investigation_data.get("completed_at", "N/A"),
                investigation_data.get("duration_minutes", "N/A"),
            ],
        }
        dataframes["Resumo"] = pd.DataFrame(summary_data)

        # Anomalies sheet
        anomalies = investigation_data.get("anomalies", [])
        if anomalies:
            anomaly_df = pd.DataFrame(anomalies)
            dataframes["Anomalias"] = anomaly_df

        # Contracts sheet
        contracts = investigation_data.get("contracts", [])
        if contracts:
            contract_df = pd.DataFrame(contracts)
            dataframes["Contratos"] = contract_df

        # Analysis results
        results = investigation_data.get("results", {})
        if results:
            results_data = []
            for key, value in results.items():
                results_data.append({"Métrica": key, "Valor": str(value)})
            dataframes["Resultados"] = pd.DataFrame(results_data)

        # Generate Excel
        return await self.generate_excel(
            data=dataframes,
            title=f"Investigação {investigation_data.get('id', 'N/A')}",
            metadata={
                "generated_at": datetime.now().isoformat(),
                "investigation_id": investigation_data.get("id", "N/A"),
            },
        )


# Global instance
export_service = ExportService()
