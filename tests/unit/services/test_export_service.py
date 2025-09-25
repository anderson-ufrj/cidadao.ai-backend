"""
Module: tests.unit.services.test_export_service
Description: Unit tests for export service
Author: Anderson H. Silva
Date: 2025-01-25
License: Proprietary - All rights reserved
"""

import pytest
import base64
from datetime import datetime
from unittest.mock import AsyncMock, MagicMock, patch
import pandas as pd
from io import BytesIO

from src.services.export_service import ExportService, export_service


class TestExportService:
    """Test suite for ExportService."""
    
    @pytest.fixture
    def service(self):
        """Create export service instance."""
        return ExportService()
    
    @pytest.fixture
    def sample_markdown_content(self):
        """Sample markdown content for testing."""
        return """
# Relatório de Investigação

## Resumo Executivo

Este relatório apresenta os resultados da investigação realizada.

## Anomalias Detectadas

### Anomalia 1
- **Tipo**: Valor atípico
- **Severidade**: 0.85
- **Descrição**: Contrato com valor 300% acima da média

### Anomalia 2
- **Tipo**: Padrão temporal
- **Severidade**: 0.72
- **Descrição**: Concentração anormal de contratos em período específico

## Conclusões

As anomalias detectadas indicam possíveis irregularidades que requerem investigação adicional.

| Métrica | Valor |
|---------|--------|
| Total de contratos | 150 |
| Valor total | R$ 1.500.000 |
| Anomalias detectadas | 12 |
"""
    
    @pytest.fixture
    def sample_dataframe(self):
        """Sample DataFrame for testing."""
        return pd.DataFrame({
            'contract_id': ['C001', 'C002', 'C003'],
            'value': [100000, 250000, 180000],
            'date': ['2024-01-15', '2024-02-20', '2024-03-10'],
            'supplier': ['Empresa A', 'Empresa B', 'Empresa C'],
            'status': ['active', 'completed', 'active']
        })
    
    @pytest.mark.asyncio
    async def test_generate_pdf_basic(self, service, sample_markdown_content):
        """Test basic PDF generation."""
        # Generate PDF
        pdf_bytes = await service.generate_pdf(
            content=sample_markdown_content,
            title="Test Report",
            metadata={'author': 'Test'},
            format_type="report"
        )
        
        # Verify PDF generated
        assert isinstance(pdf_bytes, bytes)
        assert len(pdf_bytes) > 1000  # Reasonable PDF size
        assert pdf_bytes.startswith(b'%PDF')  # PDF header
    
    @pytest.mark.asyncio
    async def test_generate_pdf_with_metadata(self, service, sample_markdown_content):
        """Test PDF generation with metadata."""
        metadata = {
            'generated_at': datetime.now().isoformat(),
            'report_type': 'investigation',
            'author': 'Cidadão.AI System'
        }
        
        pdf_bytes = await service.generate_pdf(
            content=sample_markdown_content,
            title="Investigation Report",
            metadata=metadata,
            format_type="investigation"
        )
        
        assert isinstance(pdf_bytes, bytes)
        assert pdf_bytes.startswith(b'%PDF')
    
    @pytest.mark.asyncio
    async def test_generate_excel_single_sheet(self, service, sample_dataframe):
        """Test Excel generation with single sheet."""
        excel_bytes = await service.generate_excel(
            data=sample_dataframe,
            title="Contracts Report",
            metadata={'exported_at': datetime.now().isoformat()}
        )
        
        # Verify Excel generated
        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 500  # Reasonable Excel size
        
        # Verify can be read as Excel
        df_read = pd.read_excel(BytesIO(excel_bytes), sheet_name='Dados')
        assert len(df_read) == 3
        assert 'contract_id' in df_read.columns
    
    @pytest.mark.asyncio
    async def test_generate_excel_multiple_sheets(self, service, sample_dataframe):
        """Test Excel generation with multiple sheets."""
        data = {
            'Contratos': sample_dataframe,
            'Resumo': pd.DataFrame({
                'Métrica': ['Total Contratos', 'Valor Total', 'Média'],
                'Valor': [3, 530000, 176666.67]
            })
        }
        
        excel_bytes = await service.generate_excel(
            data=data,
            title="Complete Report",
            metadata={'version': '1.0'}
        )
        
        assert isinstance(excel_bytes, bytes)
        
        # Verify sheets exist
        excel_file = pd.ExcelFile(BytesIO(excel_bytes))
        assert 'Contratos' in excel_file.sheet_names
        assert 'Resumo' in excel_file.sheet_names
        assert 'Metadados' in excel_file.sheet_names
    
    @pytest.mark.asyncio
    async def test_generate_csv(self, service, sample_dataframe):
        """Test CSV generation."""
        csv_bytes = await service.generate_csv(sample_dataframe)
        
        assert isinstance(csv_bytes, bytes)
        csv_str = csv_bytes.decode('utf-8')
        
        # Verify CSV content
        assert 'contract_id,value,date,supplier,status' in csv_str
        assert 'C001,100000' in csv_str
        assert 'Empresa A' in csv_str
    
    @pytest.mark.asyncio
    async def test_generate_bulk_export(self, service, sample_markdown_content, sample_dataframe):
        """Test bulk export with ZIP."""
        exports = [
            {
                'filename': 'report1.pdf',
                'content': sample_markdown_content,
                'format': 'pdf',
                'title': 'Report 1',
                'metadata': {}
            },
            {
                'filename': 'data.csv',
                'data': sample_dataframe,
                'format': 'csv'
            },
            {
                'filename': 'summary.txt',
                'content': 'Summary of investigations',
                'format': 'txt'
            }
        ]
        
        zip_bytes = await service.generate_bulk_export(exports)
        
        assert isinstance(zip_bytes, bytes)
        assert len(zip_bytes) > 1000  # Reasonable ZIP size
        
        # Verify ZIP content
        import zipfile
        with zipfile.ZipFile(BytesIO(zip_bytes), 'r') as zf:
            assert 'report1.pdf' in zf.namelist()
            assert 'data.csv' in zf.namelist()
            assert 'summary.txt' in zf.namelist()
    
    @pytest.mark.asyncio
    async def test_convert_investigation_to_excel(self, service):
        """Test investigation data to Excel conversion."""
        investigation_data = {
            'id': 'INV-001',
            'type': 'contract_analysis',
            'status': 'completed',
            'created_at': '2024-01-20T10:00:00',
            'completed_at': '2024-01-20T10:30:00',
            'duration_minutes': 30,
            'anomalies': [
                {
                    'type': 'value_outlier',
                    'severity': 0.85,
                    'description': 'High value detected',
                    'contract_id': 'C001'
                },
                {
                    'type': 'temporal_pattern',
                    'severity': 0.72,
                    'description': 'Unusual timing',
                    'contract_id': 'C002'
                }
            ],
            'contracts': [
                {
                    'id': 'C001',
                    'value': 500000,
                    'supplier': 'Company A'
                },
                {
                    'id': 'C002',
                    'value': 300000,
                    'supplier': 'Company B'
                }
            ],
            'results': {
                'total_analyzed': 100,
                'anomalies_found': 2,
                'risk_score': 0.78
            }
        }
        
        excel_bytes = await service.convert_investigation_to_excel(investigation_data)
        
        assert isinstance(excel_bytes, bytes)
        
        # Verify sheets
        excel_file = pd.ExcelFile(BytesIO(excel_bytes))
        assert 'Resumo' in excel_file.sheet_names
        assert 'Anomalias' in excel_file.sheet_names
        assert 'Contratos' in excel_file.sheet_names
        assert 'Resultados' in excel_file.sheet_names
    
    @pytest.mark.asyncio
    async def test_pdf_generation_thread_safety(self, service, sample_markdown_content):
        """Test PDF generation is thread-safe."""
        # Generate multiple PDFs concurrently
        import asyncio
        
        tasks = []
        for i in range(5):
            task = service.generate_pdf(
                content=sample_markdown_content + f"\n\n## Report {i}",
                title=f"Report {i}",
                metadata={'report_id': i}
            )
            tasks.append(task)
        
        results = await asyncio.gather(*tasks)
        
        # All PDFs should be generated successfully
        assert len(results) == 5
        for pdf_bytes in results:
            assert isinstance(pdf_bytes, bytes)
            assert pdf_bytes.startswith(b'%PDF')
    
    def test_custom_styles_creation(self, service):
        """Test custom PDF styles are created."""
        assert 'CustomTitle' in service.styles
        assert 'CustomSubtitle' in service.styles
        assert 'CustomBody' in service.styles
        assert 'CustomFooter' in service.styles
    
    @pytest.mark.asyncio
    async def test_excel_formatting(self, service, sample_dataframe):
        """Test Excel formatting is applied."""
        excel_bytes = await service.generate_excel(
            data=sample_dataframe,
            title="Formatted Report",
            metadata={'test': 'formatting'}
        )
        
        # Load and check formatting
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb['Dados']
        
        # Check title is merged
        assert ws.merged_cells
        
        # Check title cell has custom font
        title_cell = ws['A1']
        assert title_cell.font.bold
        assert title_cell.font.size == 16
    
    @pytest.mark.asyncio
    async def test_empty_dataframe_handling(self, service):
        """Test handling of empty DataFrames."""
        empty_df = pd.DataFrame()
        
        csv_bytes = await service.generate_csv(empty_df)
        assert csv_bytes == b''
        
        # Excel should still generate with headers
        excel_bytes = await service.generate_excel(
            data=empty_df,
            title="Empty Report"
        )
        assert isinstance(excel_bytes, bytes)
    
    @pytest.mark.asyncio
    async def test_large_content_handling(self, service):
        """Test handling of large content."""
        # Generate large markdown content
        large_content = "# Large Report\n\n"
        for i in range(100):
            large_content += f"## Section {i}\n"
            large_content += "This is a paragraph with some content. " * 50
            large_content += "\n\n"
        
        pdf_bytes = await service.generate_pdf(
            content=large_content,
            title="Large Report",
            format_type="report"
        )
        
        assert isinstance(pdf_bytes, bytes)
        assert len(pdf_bytes) > 10000  # Should be a sizeable PDF
    
    @pytest.mark.asyncio
    async def test_special_characters_handling(self, service):
        """Test handling of special characters."""
        content_with_special = """
# Relatório com Caracteres Especiais

## Seção com acentuação: áéíóú àèìòù ãõ ç

### Símbolos: @#$%^&*()_+-={}[]|:";'<>?,./

**Texto em negrito** e *texto em itálico*

Código: `print("Olá, Mundo!")`
"""
        
        pdf_bytes = await service.generate_pdf(
            content=content_with_special,
            title="Relatório Especial"
        )
        
        assert isinstance(pdf_bytes, bytes)
        assert pdf_bytes.startswith(b'%PDF')
    
    def test_global_service_instance(self):
        """Test global service instance is available."""
        assert export_service is not None
        assert isinstance(export_service, ExportService)